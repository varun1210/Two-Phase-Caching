import random;
from openpyxl import Workbook;
from openpyxl import load_workbook;

cf5Calculator = [0] * 50;
cf10Calculator = [0] * 100;
wb = Workbook();
wb = load_workbook('D:\VIT Semesters\Fall Semester 2020-21\Projects\Storage\Storage Dataset.xlsx')
ws = wb.active;
for i in range(2, 2002):
    ws['A' + str(i)].value = i - 1;

for i in range(2, 2002):
    ws['B' + str(i)].value = random.randint(1, 50);

for i in range(2, 2002):
    currVal = ws['B' + str(i)].value;
    ws['C' + str(i)].value = cf5Calculator[currVal - 1] + 1;
    cf5Calculator[currVal - 1] += 1;
    ws['D' + str(i)].value = cf10Calculator[currVal - 1] + 1;
    cf10Calculator[currVal - 1] += 1;
    if((i - 1) % 5 == 0):
        cf5Calculator = [0] * 50;
    if((i - 1) % 10 == 0):
        cf10Calculator = [0] * 100;
wb.save('D:\VIT Semesters\Fall Semester 2020-21\Projects\Storage\Storage Dataset.xlsx');










