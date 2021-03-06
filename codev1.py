import pandas as pd;
from pandas import DataFrame as df;
import random;
from openpyxl import Workbook;
from openpyxl import load_workbook;


#The two lists maintain the status of the short term and long term cache
shortTermCache = set();
shortTermCacheCapacity = 5;
longTermCache = set();
longTermCacheCapacity = 2 * shortTermCacheCapacity;
registerData = {};
registerSize = 100;


def fillRegisterStatus(t, cache):
    print("Register resetting...");
    regRequiredData = cache.loc[(cache["T"] > t) & (cache["T"] <= t + registerSize), ["T", "P"]];
    for record in regRequiredData["P"]:
        if(record not in registerData):
            registerData[record] = 1;
        else:
            registerData[record] = registerData[record] + 1;         
    print("Reset complete. Register values after reset are:\n" + str(registerData) + "\n");
    overflow = fillShortTermCache();
    fillLongTermCache(regRequiredData, overflow);
    print(longTermCache);

"""
def fillShortTermCache():
    print("Short term cache resetting...");
    cf = 0;
    overflowList = [];
    fillCount = 0;
    currentRegisterSize = len(registerData);
    for process in registerData:
        cf += registerData[process];
    averageFrequency = cf / currentRegisterSize;
    print("The threshold frequency thus calculated is :" + str(averageFrequency));
    for process in registerData:
        if(registerData[process] > averageFrequency):
            i = 0;
            while((i < len(shortTermCache)) and (shortTermCache[i] != -1)):
                i += 1;
            fillCount += 1;
            if(i >= len(shortTermCache)):
                overflowList.append(process);
            else:
                shortTermCache[i] = process;
                
    if(fillCount > 5):
       i = 0;
       for process in overflowList:
           longTermCache[i] = process;
           i += 1;
    elif(fillCount < 5):
        #YOU ARE HERE AT THE TIME OF WRITING THIS CODE. WE ARE HANDLING THE CASE IN WHICH SHORT TERM CACHE HAS TO BE FILLED WITH
        #THE REMAINING ITEMS BASED ON THE RANK OF THE REMAINING ITEMS.
        cacheSet = 
    
    print("Short term cache has been reset. The new values are:\n" + str(shortTermCache) + "\n");
"""

def fillShortTermCache():
    print("Short term cache resetting...");
    cf = 0;
    overflowList = [];
    currentRegisterSize = len(registerData);
    for process in registerData:
        cf += registerData[process];
    averageFrequency = cf / currentRegisterSize;
    print("The threshold frequency thus calculated is :" + str(averageFrequency));
    for process in registerData:
        if(registerData[process] > averageFrequency):            
            if(len(shortTermCache) > shortTermCacheCapacity):
                overflowList.append(process);
            else:
                shortTermCache.add(process);
    if(len(shortTermCache) < shortTermCacheCapacity):
        #pseudoReg = {k : v for k, v in sorted(registerData.items(), key = lambda item : item[1])};
        pseudoReg = sorted(registerData.items(), key = lambda kv:(kv[1], kv[0]), reverse = True);
        #print("PSEUDO REGISTER DATAAAA ---> " + str(pseudoReg))
        #print("REGISTER DATAAAAA -----> " + str(registerData));
        for process in pseudoReg:
            if(len(shortTermCache) >= shortTermCacheCapacity):
                break;
            if(process[0] in shortTermCache):
               continue;
            else:
               shortTermCache.add(process[0]);
    print("The short term cache has been filled. It's elements are:\n" + str(shortTermCache));
    return overflowList;

def fillLongTermCache(regReq, overflow):
    for process in overflow:
        longTermCache.add(process);
        if(len(longTermCache) >= longTermCacheCapacity):
            print("The long term cache has been filled. It's elements are:\n" + str(longTermCache));
            return;
    processList = regReq["P"].tolist();
    #fringe creation
    shortTermFringe = [];
    for process in shortTermCache:
        if(len(longTermCache) >= longTermCacheCapacity):               
            print("The long term cache has been filled. It's elements are:\n" + str(longTermCache));
            return;
        processIndicies = [];
        for i in range(len(processList)):
            if(processList[i] == process):
               processIndicies.append(i);
        potentialCache = {};
        for i in range(len(processIndicies)):
            ##
            if(((processIndicies[i] + 2) in range(0, len(processList))) and ((processIndicies[i] - 2) in range(0, len(processList)))):
                x1 = registerData[processList[processIndicies[i] + 1]];
                x2 = registerData[processList[processIndicies[i] + 2]];
                w1 = 1;
                w2 = -0.75;
                yin = (x1 * w1) + (x2 * w2);
                if(yin > 0):
                    potentialCache[processList[processIndicies[i] + 1]] = registerData[processList[processIndicies[i] + 1]];
                else:
                    potentialCache[processList[processIndicies[i] + 2]] = registerData[processList[processIndicies[i] + 2]];
            potentialCache = {k : v for k, v in sorted(potentialCache.items(), key = lambda item : item[1])};
            count = 0;
            for frequency in potentialCache:
               if(count >= 3):
                   break;
               if(potentialCache[frequency] not in shortTermCache):
                   longTermCache.add(potentialCache[frequency]);
    if(len(longTermCache) < longTermCacheCapacity):
        #pseudoReg = {k : v for k, v in sorted(registerData.items(), key = lambda item : item[1])};
        pseudoReg = sorted(registerData.items(), key = lambda kv:(kv[1], kv[0]), reverse = True);
        print("PSEUDO REGISTER DATAAAA ---> " + str(pseudoReg))
        print("REGISTER DATAAAAA -----> " + str(registerData));
        for process in pseudoReg:
            if(len(longTermCache) >= longTermCacheCapacity):
                print("The long term cache has been filled. It's elements are:\n" + str(longTermCache));
                return;               
            if(process[0] not in shortTermCache):
                longTermCache.add(process[0]);
               
        
               
                
               
                   
               
                       
      
                    
                     
"""
cf5Calculator = [0, 0, 0, 0, 0, 0, 0, 0];
cf10Calculator = [0, 0, 0, 0, 0, 0, 0, 0];
wb = Workbook();
wb = load_workbook('D:\VIT Semesters\Fall Semester 2020-21\Projects\Storage\Storage Dataset.xlsx')
ws = wb.active;
for i in range(2, 2002):
    ws['A' + str(i)].value = i - 1;

for i in range(2, 2002):
    ws['B' + str(i)].value = random.randint(1, 8);

for i in range(2, 2002):
    currVal = ws['B' + str(i)].value;
    ws['C' + str(i)].value = cf5Calculator[currVal - 1] + 1;
    cf5Calculator[currVal - 1] += 1;
    ws['D' + str(i)].value = cf10Calculator[currVal - 1] + 1;
    cf10Calculator[currVal - 1] += 1;
    if((i - 1) % 5 == 0):
        cf5Calculator = [0, 0, 0, 0, 0, 0, 0, 0];
    if((i - 1) % 10 == 0):
        cf10Calculator = [0, 0, 0, 0, 0, 0, 0, 0];
wb.save('D:\VIT Semesters\Fall Semester 2020-21\Projects\Storage\Storage Dataset.xlsx');
"""


cacheData = pd.read_excel("D:\VIT Semesters\Fall Semester 2020-21\Projects\Storage\Storage Dataset.xlsx");
fillRegisterStatus(90, cacheData);






